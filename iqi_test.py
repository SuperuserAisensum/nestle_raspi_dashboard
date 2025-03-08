import cv2
import numpy as np
import math
import os
import pandas as pd
from glob import glob
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
import io
from PIL import Image as PILImage
from sklearn.cluster import KMeans

### PART A: DARKNESS INDEX ###
def compute_darkness_index(image):
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    avg_intensity = np.mean(gray_image)
    darkness_index = (avg_intensity / 255) * 100
    if darkness_index <= 35:
        score = (darkness_index / 35) * 100
    elif 35 < darkness_index <= 55:
        score = ((darkness_index - 35) / 20) * 100
        score = 50 + (score / 2)
    elif 55 < darkness_index <= 75:
        score = ((75 - darkness_index) / 20) * 100
        score = 50 + (score / 2)
    else:
        score = ((100 - darkness_index) / 25) * 100
    return round(max(0, min(100, score)), 2)

### PART B: ANGLE OF SHOT INDEX ###
def compute_intersection(line1, line2, image_shape):
    x1, y1, x2, y2 = line1
    x3, y3, x4, y4 = line2
    h, w = image_shape

    # Normalize coordinates to [0, 1] range
    x1, y1 = x1 / w, y1 / h
    x2, y2 = x2 / w, y2 / h
    x3, y3 = x3 / w, y3 / h
    x4, y4 = x4 / w, y4 / h

    A1, B1, C1 = y2 - y1, x1 - x2, (y2 - y1) * x1 + (x1 - x2) * y1
    A2, B2, C2 = y4 - y3, x3 - x4, (y4 - y3) * x3 + (x3 - x4) * y3
    determinant = A1 * B2 - A2 * B1

    # Check for near-parallel lines or overflow
    if abs(determinant) < 1e-10:  # Threshold for near-zero determinant
        return None

    x = (B2 * C1 - B1 * C2) / determinant
    y = (A1 * C2 - A2 * C1) / determinant

    # Denormalize back to pixel coordinates
    x, y = x * w, y * h
    return (x, y)

def find_vanishing_point(image, image_path):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    gray = cv2.GaussianBlur(gray, (5, 5), 0)
    edges = cv2.Canny(gray, 50, 150, apertureSize=3)

    lines = cv2.HoughLinesP(edges, 1, np.pi / 180, threshold=100, 
                           minLineLength=150, maxLineGap=10)
    h, w = image.shape[:2]
    cx, cy = w / 2.0, h / 2.0

    vp = (cx, cy)
    filtered_lines = []
    intersections = []

    if lines is not None and len(lines) >= 2:
        for line in lines:
            x1, y1, x2, y2 = line[0]
            length = math.sqrt((x2 - x1)**2 + (y2 - y1)**2)
            angle = abs(math.degrees(math.atan2(y2 - y1, x2 - x1)))
            if (15 < angle < 75 or 105 < angle < 165) and length > 200:
                filtered_lines.append(line[0])

        # print(f"Debug: Found {len(filtered_lines)} diagonal lines")
        if len(filtered_lines) >= 2:
            margin = w / 2
            center_threshold = w / 3
            for i in range(len(filtered_lines)):
                for j in range(i + 1, len(filtered_lines)):
                    pt = compute_intersection(filtered_lines[i], filtered_lines[j], (h, w))
                    if pt and -margin <= pt[0] <= w + margin and -margin <= pt[1] <= h + margin:
                        distance_to_center = math.sqrt((pt[0] - cx)**2 + (pt[1] - cy)**2)
                        if distance_to_center < center_threshold:
                            intersections.append(pt)

            # print(f"Debug: Found {len(intersections)} valid intersections")
            if intersections:
                intersections = np.array(intersections)
                if len(intersections) > 3:
                    try:
                        kmeans = KMeans(n_clusters=min(3, len(intersections)), n_init=10).fit(intersections)
                        centers = kmeans.cluster_centers_
                        distances = [math.sqrt((x - cx)**2 + (y - cy)**2) for x, y in centers]
                        best_cluster_idx = np.argmin(distances)
                        vp = tuple(centers[best_cluster_idx])
                    except Exception as e:
                        # print(f"Debug: Clustering failed - {str(e)}, using weighted mean")
                        weights = [1 / (math.sqrt((x - cx)**2 + (y - cy)**2) + 1) for x, y in intersections]
                        vp = tuple(np.average(intersections, axis=0, weights=weights))
                else:
                    vp = tuple(np.median(intersections, axis=0))
            else:
                # print("Debug: No central intersections, assuming centered perspective")
                pass
        else:
            # print("Debug: Too few lines, assuming centered perspective")
            pass
    else:
        # print(f"Debug: Found {0 if lines is None else len(lines)} lines - need at least 2")
        pass

    # print(f"Debug: Vanishing Point at {vp}")

    # --- BEGIN HASHED: Visualization Section ---
    # vis_image = image.copy()
    # marker_scale = w // 50
    # for line in filtered_lines:
    #     x1, y1, x2, y2 = line
    #     cv2.line(vis_image, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), max(2, marker_scale // 5))
    # for x, y in intersections:
    #     cv2.circle(vis_image, (int(x), int(y)), max(10, marker_scale // 2), (0, 0, 255), -1)
    # cv2.circle(vis_image, (int(vp[0]), int(vp[1])), max(15, marker_scale), (255, 255, 255), max(4, marker_scale // 5))
    # cv2.circle(vis_image, (int(vp[0]), int(vp[1])), max(15, marker_scale), (255, 0, 0), -1)
    # cross_size = max(20, marker_scale)
    # thickness = max(3, marker_scale // 5)
    # cv2.line(vis_image, (int(cx), int(cy - cross_size)), (int(cx), int(cy + cross_size)), (0, 255, 255), thickness)
    # cv2.line(vis_image, (int(cx - cross_size), int(cy)), (int(cx + cross_size), int(cy)), (0, 255, 255), thickness)
    #
    # debug_dir = os.path.join(os.path.dirname(image_path), 'debug_visuals')
    # os.makedirs(debug_dir, exist_ok=True)
    # vis_path = os.path.join(debug_dir, f"debug_{os.path.basename(image_path)}")
    # cv2.imwrite(vis_path, vis_image)
    # print(f"Debug: Visualization saved to {vis_path}")
    # --- END HASHED: Visualization Section ---

    return vp

def compute_angle_from_vanishing_point(vp, image, focal_length=1200):
    h, w = image.shape[:2]
    cx, cy = w / 2.0, h / 2.0
    dx = vp[0] - cx
    dy = vp[1] - cy
    distance = math.sqrt(dx**2 + dy**2)
    angle_rad = math.atan2(distance, focal_length)
    angle_deg = math.degrees(angle_rad)
    return angle_deg

def compute_angle_score(angle):
    if angle is None:
        return 50
    score = 100 * math.exp(-angle / 15)
    return round(max(0, min(100, score)), 2)

### FINAL IQI COMPUTATION ###
def compute_IQI(image_path, focal_length=1200):
    image = cv2.imread(image_path)
    if image is None:
        print(f"Error: Image {image_path} not found or cannot be loaded.")
        return None

    darkness_index = compute_darkness_index(image)
    vp = find_vanishing_point(image, image_path)
    # The following condition checks if the vanishing point remains at image center.
    if vp == (image.shape[1] / 2.0, image.shape[0] / 2.0) and not any(find_vanishing_point(image, image_path)):
        # print(f"Warning: Could not detect a vanishing point for {image_path}. Using default angle index.")
        angle = None
        angle_index = 50
    else:
        angle = compute_angle_from_vanishing_point(vp, image, focal_length)
        angle_index = compute_angle_score(angle)

    if angle is None:
        IQI = round((darkness_index * 0.7) + (angle_index * 0.3), 2)
    else:
        IQI = round((darkness_index * 0.5) + (angle_index * 0.5), 2)

    return {
        'image_path': image_path,
        'darkness_index': darkness_index,
        'angle_index': angle_index,
        'angle': angle,
        'iqi_score': IQI
    }

if __name__ == "__main__":
    imgs_dir = r'images'
    folder_name = os.path.basename(os.path.normpath(imgs_dir))

    if not os.path.exists(imgs_dir):
        print(f"Error: Directory '{imgs_dir}' does not exist!")
        exit(1)

    image_files = glob(os.path.join(imgs_dir, '*'))
    if not image_files:
        print(f"Error: No images found in '{imgs_dir}' directory!")
        exit(1)
        
    # print(f"Found {len(image_files)} images to process...")
    results = []
    
    wb = Workbook()
    ws = wb.active
    ws.title = "IQI Results"
    headers = ['Image Path', 'Darkness Index', 'Angle Index', 'Angle', 'IQI Score', 'Image']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    current_row = 2
    
    for image_path in image_files:
        # print(f"\nProcessing: {os.path.basename(image_path)}")
        result = compute_IQI(image_path)
        if result:
            ws.cell(row=current_row, column=1, value=result['image_path'])
            ws.cell(row=current_row, column=2, value=result['darkness_index'])
            ws.cell(row=current_row, column=3, value=result['angle_index'])
            ws.cell(row=current_row, column=4, value=result['angle'])
            ws.cell(row=current_row, column=5, value=result['iqi_score'])
            
            try:
                img = PILImage.open(image_path)
                max_size = (100, 100)
                img.thumbnail(max_size)
                img_byte_arr = io.BytesIO()
                img.save(img_byte_arr, format='PNG')
                img_byte_arr.seek(0)
                img = XLImage(img_byte_arr)
                ws.add_image(img, f'F{current_row}')
                ws.row_dimensions[current_row].height = 75
            except Exception as e:
                print(f"Error adding image to Excel: {str(e)}")
            
            current_row += 1
            results.append(result)
            
            # print(f"Results for {os.path.basename(image_path)}:")
            # print(f"Darkness Index: {result['darkness_index']}/100")
            # print(f"Angle Score: {result['angle_index']}/100")
            # if result['angle']:
            #     print(f"Angle: {result['angle']:.2f}°")
            # print(f"Final IQI Score: {result['iqi_score']}/100")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    output_file = f'iqi_results_{folder_name}.xlsx'
    try:
        wb.save(output_file)
        # print(f"\nResults with images successfully saved to {os.path.abspath(output_file)}")
    except Exception as e:
        print(f"\nError saving Excel file: {str(e)}")
        try:
            df = pd.DataFrame(results)
            df.to_excel(output_file, index=False)
            # print(f"Saved results without images to {os.path.abspath(output_file)}")
        except Exception as e:
            print(f"Error saving results: {str(e)}")
